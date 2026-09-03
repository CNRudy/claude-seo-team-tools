#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
推广合作方体检工具 (Partner Health Checker)
============================================
基于 claude-seo 的抓取/解析/内容质量能力，对推广合作方网站做多维体检：

  1. 运营状态    - 可访问性、HTTP 状态、响应时间
  2. 技术 SEO    - HTTPS 强制、robots.txt、sitemap、title/desc/canonical、移动适配
  3. 内容质量    - 正文质量评分(AI味/填充/信息密度)、正文长度、E-E-A-T 信号页
  4. 外链画像    - (可选) Common Crawl PageRank / 中心度，无需 API key
  5. 综合评分    - 加权合成 0-100 分 + A/B/C/D 分级

输出：Excel (.xlsx) + CSV 体检报告。

用法:
    python partner_health_check.py --input urls.txt
    python partner_health_check.py --urls https://example.com https://example.org
    python partner_health_check.py --input urls.txt --pagerank --workers 4
"""

from __future__ import annotations

import argparse
import concurrent.futures as cf
import csv
import json
import os
import re
import subprocess
import sys
import time
from datetime import datetime

# ---------------------------------------------------------------- paths
_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_SCRIPTS = os.path.join(_REPO_ROOT, "claude-seo", "scripts")
if _SCRIPTS not in sys.path:
    sys.path.insert(0, _SCRIPTS)

try:
    import requests
except ImportError:
    print("Error: requests required", file=sys.stderr)
    sys.exit(1)

try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

try:
    import trafilatura
except ImportError:
    trafilatura = None

try:
    import htmldate
except ImportError:
    htmldate = None

# 复用 claude-seo 脚本
from batch_site_analyzer import _fetch_compat, normalize_url  # noqa: E402

try:
    from content_quality import analyse as content_quality_analyse  # noqa: E402
except Exception:
    content_quality_analyse = None

try:
    from parse_html import parse_html  # noqa: E402
except Exception:
    parse_html = None

_PY = sys.executable


# ---------------------------------------------------------------- helpers
def _http_get(url: str, timeout: int = 12) -> dict:
    """Minimal GET with UA; returns {status, text, headers, final_url, error}."""
    out = {"status": None, "text": "", "headers": {}, "final_url": url, "error": None}
    try:
        r = requests.get(
            url, timeout=timeout, allow_redirects=True,
            headers={
                "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                               "AppleWebKit/537.36 (KHTML, like Gecko) "
                               "Chrome/126.0 Safari/537.36"),
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            },
        )
        out["status"] = r.status_code
        out["headers"] = dict(r.headers)
        out["final_url"] = r.url
        out["text"] = r.text
    except requests.exceptions.RequestException as e:
        out["error"] = str(e)[:160]
    return out


def _check_robots_txt(domain: str, timeout: int = 12) -> dict:
    """Fetch robots.txt; return dict with existence, disallow-all, blocked."""
    base = f"https://{domain}/robots.txt"
    r = _http_get(base, timeout)
    if r["error"] or r["status"] is None:
        return {"exists": False, "blocked": False, "note": "获取失败"}
    if r["status"] >= 400:
        return {"exists": False, "blocked": False, "note": f"HTTP {r['status']}"}
    body = r["text"].lower()
    # 只看 User-agent: * 块内的规则；特定 UA（如 AI bot）的屏蔽不算全站屏蔽
    blocked_all = False
    ua_blocks = re.split(r"(?=^user-agent:)", body, flags=re.MULTILINE)
    for block in ua_blocks:
        if block.startswith("user-agent: *"):
            blocked_all = bool(re.search(r"^disallow:\s*/\s*$", block, re.MULTILINE))
            break
    return {
        "exists": True,
        "blocked": blocked_all,
        "note": "正常" if body.strip() else "空文件",
    }


def _check_sitemap(domain: str, timeout: int = 12) -> dict:
    """Try common sitemap locations."""
    for path in ("/sitemap.xml", "/sitemap_index.xml", "/sitemap/sitemap.xml"):
        r = _http_get(f"https://{domain}{path}", timeout=8)
        if r["error"]:
            continue
        if r["status"] == 200 and ("<urlset" in r["text"][:4000].lower()
                                   or "<sitemapindex" in r["text"][:4000].lower()):
            return {"found": True, "path": path, "status": 200}
        if r["status"] == 200:
            return {"found": True, "path": path, "status": 200}
    return {"found": False, "path": "", "status": None}


def _check_trust_pages(parsed: dict, body_text: str) -> dict:
    """E-E-A-T signal: about/contact/privacy pages linked from homepage, or
    discoverable at conventional paths."""
    links = []
    try:
        links = [l.get("href", "") for l in (parsed.get("links") or {}).get("internal", [])]
    except Exception:
        pass
    joined = " ".join(links).lower() + " " + body_text.lower()
    about = bool(re.search(r"about|about-us|about_us|关于", joined))
    contact = bool(re.search(r"contact|contact-us|联系我们|联系方式", joined))
    privacy = bool(re.search(r"privacy|隐私", joined))
    return {
        "about": about,
        "contact": contact,
        "privacy": privacy,
        "terms": bool(re.search(r"terms|条款", joined)),
    }


def _probe_trust_paths(domain: str, timeout: int = 6) -> dict:
    """Probe conventional trust-page paths directly (robust to bot walls)."""
    paths = {
        "about": ("/about", "/about-us", "/aboutus", "/about/"),
        "contact": ("/contact", "/contact-us", "/contactus", "/contact/"),
        "privacy": ("/privacy", "/privacy-policy", "/privacy/"),
    }
    found = {}
    for key, cands in paths.items():
        ok = False
        for p in cands:
            r = _http_get(f"https://{domain}{p}", timeout=timeout)
            if r["error"]:
                continue
            if r["status"] and r["status"] < 400 and len(r["text"]) > 400:
                ok = True
                break
        found[key] = ok
    return found


def _extract_publication_date(raw_html: str, text: str) -> str:
    """Try htmldate then regex on the raw HTML."""
    if htmldate is not None:
        try:
            d = htmldate.find_date(raw_html)
            if d:
                return str(d)[:10]
        except Exception:
            pass
    m = re.search(r"(20\d{2})[-/.](\d{1,2})[-/.](\d{1,2})", text[:6000])
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return ""


def _pagerank_info(domain: str) -> dict:
    """Optional Common Crawl PageRank query (slow; downloads index on first run)."""
    try:
        proc = subprocess.run(
            [_PY, os.path.join(_SCRIPTS, "commoncrawl_graph.py"), domain, "--json"],
            capture_output=True, text=True, timeout=180,
        )
        if proc.returncode == 0:
            data = json.loads(proc.stdout)
            d = data.get("data", {})
            return {
                "in_crawl": d.get("in_crawl"),
                "pagerank": d.get("pagerank"),
                "harmonic_centrality": d.get("harmonic_centrality"),
                "n_hosts": d.get("n_hosts"),
                "note": (d.get("note") or "")[:60],
            }
    except Exception as e:
        return {"error": str(e)[:80]}
    return {"error": "查询失败"}


# ---------------------------------------------------------------- scoring
def _score(domain: str, status: str, tech: dict, content: dict) -> dict:
    """Composite 0-100 score with weights: 运营20 / 技术35 / 内容45."""
    s = {"运营": 0.0, "技术": 0.0, "内容": 0.0, "总分": 0.0, "分级": "D"}

    # 运营 (20)
    op = 0.0
    if status == "有效":
        op = 20.0
    elif status == "受限":
        op = 12.0
    s["运营"] = op

    # 技术 (35)
    t = 0.0
    if tech.get("https_ok"):
        t += 6
    if tech.get("robots_ok"):
        t += 6
    if tech.get("sitemap"):
        t += 6
    if tech.get("title"):
        t += 6
    if tech.get("meta_desc"):
        t += 4
    if tech.get("viewport"):
        t += 3
    if tech.get("canonical"):
        t += 2
    if tech.get("h1"):
        t += 2
    s["技术"] = min(t, 35.0)

    # 内容 (45)
    c = 0.0
    cq = content.get("quality_score")
    if cq is not None:
        c += min(18.0, cq / 100.0 * 18.0)  # 质量分最高 18
    wc = content.get("word_count") or 0
    if wc >= 800:
        c += 9
    elif wc >= 300:
        c += 6
    elif wc >= 100:
        c += 3
    trust = content.get("trust") or {}
    for k in ("about", "contact", "privacy"):
        if trust.get(k):
            c += 4
    if content.get("pub_date"):
        c += 2
    s["内容"] = min(c, 45.0)

    total = s["运营"] + s["技术"] + s["内容"]
    grade = "A" if total >= 85 else ("B" if total >= 70 else ("C" if total >= 55 else "D"))
    s["总分"] = round(total, 1)
    s["分级"] = grade
    return s


# ---------------------------------------------------------------- per-site
def health_check(url: str, timeout: int = 15, with_pagerank: bool = False) -> dict:
    """Run the full health check for one site. Returns a result row dict."""
    url = normalize_url(url)
    domain = re.sub(r"^www\.", "", url.split("//")[-1].split("/")[0]).lower() or url
    row = {
        "URL": url,
        "域名": domain,
        "状态": "无效",
        "综合评分": "",
        "分级": "D",
        "HTTP状态码": "",
        "响应时间(秒)": "",
        "最终URL": "",
        "HTTPS": "否",
        "robots.txt": "",
        "sitemap": "未发现",
        "标题": "无",
        "Meta描述": "无",
        "移动适配": "否",
        "Canonical": "无",
        "正文质量分": "",
        "AI/填充信号": "",
        "正文词数": 0,
        "正文摘要": "",
        "关于页": "否",
        "联系页": "否",
        "隐私页": "否",
        "最后更新": "",
        "PageRank": "",
        "外链主机数": "",
        "体检说明": "",
    }

    if not url:
        row["体检说明"] = "空 URL"
        return row

    # 1. 抓主页
    t0 = time.time()
    result = _fetch_compat(url, timeout=timeout)
    elapsed = round(time.time() - t0, 2)
    row["响应时间(秒)"] = elapsed
    final_url = result.get("url") or url
    row["最终URL"] = final_url
    status = result.get("status_code")

    if result.get("error"):
        row["体检说明"] = result["error"][:120]
        row["综合评分"] = 0
        return row
    if status is None or status >= 400:
        row["HTTP状态码"] = str(status or "")
        if status in (401, 403, 429):
            row["状态"] = "受限"
            row["体检说明"] = f"HTTP {status} 访问受限"
        else:
            row["体检说明"] = f"HTTP {status}"
        row["综合评分"] = 12 if row["状态"] == "受限" else 0
        row["分级"] = "D"
        return row

    row["HTTP状态码"] = str(status)
    content = result.get("content") or ""

    # HTTPS 强制跳转判定
    row["HTTPS"] = "是" if final_url.lower().startswith("https://") else "否"

    # 解析
    parsed = {}
    body_text = ""
    try:
        if parse_html is not None:
            parsed = parse_html(content)
        if trafilatura is not None:
            body_text = trafilatura.extract(content, include_comments=False) or ""
    except Exception:
        pass

    title = (parsed.get("title") or "").strip()
    meta_desc = (parsed.get("meta_description") or "").strip()
    canonical = (parsed.get("canonical") or "").strip()
    h1s = parsed.get("h1") or []
    word_count = parsed.get("word_count") or len(body_text.split())
    has_viewport = "viewport" in content.lower()[:6000]

    row["状态"] = "有效" if (title or word_count > 0) else "无效"
    if row["状态"] == "无效":
        row["体检说明"] = "可访问但无有效内容"
        row["综合评分"] = 0
        return row

    row["标题"] = "有" if title else "无"
    row["Meta描述"] = "有" if meta_desc else "无"
    row["移动适配"] = "是" if has_viewport else "否"
    row["Canonical"] = "有" if canonical else "无"
    row["正文词数"] = word_count
    row["正文摘要"] = re.sub(r"\s+", " ", body_text[:120]).strip()

    # 2. 技术子检查
    robots = _check_robots_txt(domain, timeout=8)
    sitemap = _check_sitemap(domain, timeout=8)
    row["robots.txt"] = ("存在(未屏蔽)" if robots["exists"] and not robots["blocked"]
                         else ("存在(屏蔽抓取!)" if robots["exists"] else "无/无法获取"))
    row["sitemap"] = ("发现" if sitemap.get("found") else "未发现")
    tech = {
        "https_ok": row["HTTPS"] == "是",
        "robots_ok": robots["exists"] and not robots["blocked"],
        "sitemap": bool(sitemap.get("found")),
        "title": bool(title),
        "meta_desc": bool(meta_desc),
        "viewport": has_viewport,
        "canonical": bool(canonical),
        "h1": bool(h1s),
    }

    # 3. 内容质量
    cq = None
    flags = []
    if content_quality_analyse is not None and len(body_text.split()) >= 30:
        try:
            cq_res = content_quality_analyse(body_text)
            cq = cq_res.get("overall_quality")
            flags = cq_res.get("flags") or []
        except Exception:
            pass
    row["正文质量分"] = cq if cq is not None else ""
    row["AI/填充信号"] = "、".join(flags[:4]) if flags else "无明显信号"

    trust_link = _check_trust_pages(parsed, body_text)
    trust_path = _probe_trust_paths(domain, timeout=6)
    trust = {
        "about": trust_link.get("about") or trust_path.get("about"),
        "contact": trust_link.get("contact") or trust_path.get("contact"),
        "privacy": trust_link.get("privacy") or trust_path.get("privacy"),
        "terms": trust_link.get("terms"),
    }
    row["关于页"] = "是" if trust.get("about") else "否"
    row["联系页"] = "是" if trust.get("contact") else "否"
    row["隐私页"] = "是" if trust.get("privacy") else "否"
    row["最后更新"] = _extract_publication_date(content, body_text)

    content_metrics = {
        "quality_score": cq,
        "word_count": word_count,
        "trust": trust,
        "pub_date": bool(row["最后更新"]),
    }

    # 4. 可选外链画像
    if with_pagerank:
        pr = _pagerank_info(domain)
        row["PageRank"] = str(pr.get("pagerank") or "") if pr.get("pagerank") is not None else ""
        row["外链主机数"] = str(pr.get("n_hosts") or "") if pr.get("n_hosts") is not None else ""
        if pr.get("note"):
            row["体检说明"] = pr["note"]

    # 5. 综合评分
    score = _score(domain, row["状态"], tech, content_metrics)
    row["综合评分"] = score["总分"]
    row["分级"] = score["分级"]
    if not row["体检说明"]:
        parts = []
        if not tech["robots_ok"]:
            parts.append("robots 屏蔽")
        if not tech["sitemap"]:
            parts.append("无 sitemap")
        if cq is not None and cq < 60:
            parts.append(f"内容质量偏低({cq})")
        row["体检说明"] = "、".join(parts) if parts else "整体健康"

    return row


# ---------------------------------------------------------------- batch
def run_batch(urls: list, workers: int = 4, timeout: int = 15,
              with_pagerank: bool = False) -> list:
    rows = []
    with cf.ThreadPoolExecutor(max_workers=workers) as pool:
        futs = {pool.submit(health_check, u, timeout, with_pagerank): u for u in urls}
        for i, fut in enumerate(cf.as_completed(futs), 1):
            u = futs[fut]
            try:
                row = fut.result()
            except Exception as e:
                row = {"URL": u, "状态": "无效", "综合评分": 0, "分级": "D",
                       "体检说明": f"异常: {str(e)[:100]}"}
            rows.append(row)
            print(f"[{i}/{len(urls)}] {row.get('分级','-')} {row.get('综合评分','-')}  {u}")
    order = {u: i for i, u in enumerate(urls)}
    rows.sort(key=lambda r: order.get(r.get("URL", ""), 0))
    return rows


HEADERS = [
    "URL", "域名", "状态", "综合评分", "分级", "HTTP状态码", "响应时间(秒)",
    "最终URL", "HTTPS", "robots.txt", "sitemap", "标题", "Meta描述",
    "移动适配", "Canonical", "正文质量分", "AI/填充信号", "正文词数",
    "正文摘要", "关于页", "联系页", "隐私页", "最后更新",
    "PageRank", "外链主机数", "体检说明",
]


def export_excel(rows: list, path: str) -> None:
    if not _HAS_OPENPYXL:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合作方体检报告"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    grade_fill = {
        "A": PatternFill("solid", fgColor="C6EFCE"),
        "B": PatternFill("solid", fgColor="E2EFDA"),
        "C": PatternFill("solid", fgColor="FFF2CC"),
        "D": PatternFill("solid", fgColor="FCE4EC"),
    }
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        ws.append([row.get(h, "") for h in HEADERS])
    for r in range(2, ws.max_row + 1):
        g = ws.cell(row=r, column=5).value
        fill = grade_fill.get(g)
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {"A": 40, "B": 26, "C": 8, "D": 10, "E": 6, "F": 10, "G": 11,
              "H": 42, "I": 7, "J": 16, "K": 9, "L": 6, "M": 9, "N": 9,
              "O": 9, "P": 10, "Q": 14, "R": 9, "S": 42, "T": 7, "U": 7,
              "V": 7, "W": 12, "X": 9, "Y": 10, "Z": 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
    wb.save(path)


def export_csv(rows: list, path: str) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=HEADERS, extrasaction="ignore")
        w.writeheader()
        for row in rows:
            w.writerow({h: row.get(h, "") for h in HEADERS})


def read_urls(path: str) -> list:
    out = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#"):
                out.append(line)
    return out


def main() -> int:
    ap = argparse.ArgumentParser(
        description="推广合作方体检：运营/技术/内容/外链四维评分 (0-100, A-D)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--input", help="URL 列表文件（每行一个）")
    ap.add_argument("--urls", nargs="+", help="直接传入 URL")
    ap.add_argument("--workers", type=int, default=4)
    ap.add_argument("--timeout", type=int, default=15)
    ap.add_argument("--pagerank", action="store_true", help="附加 Common Crawl 外链画像（较慢）")
    ap.add_argument("--output", help="输出 Excel 路径")
    args = ap.parse_args()

    urls = read_urls(args.input) if args.input else ([normalize_url(u) for u in (args.urls or [])])
    if not urls:
        ap.error("请提供 --input 或 --urls")
    seen, dedup = set(), []
    for u in urls:
        if u and u not in seen:
            seen.add(u)
            dedup.append(u)
    urls = dedup

    print(f"开始体检 {len(urls)} 个站点（并发 {args.workers}）...")
    rows = run_batch(urls, workers=args.workers, timeout=args.timeout,
                     with_pagerank=args.pagerank)
    from collections import Counter
    grade_cnt = Counter(r.get("分级", "D") for r in rows)
    print("\n=== 体检汇总 ===")
    print(f"总数: {len(rows)} | " + " | ".join(f"{k}: {v}" for k, v in sorted(grade_cnt.items())))

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = args.output or os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                      f"partner_health_{stamp}.xlsx")
    if not out.endswith((".xlsx", ".xls")):
        out += ".xlsx"
    if _HAS_OPENPYXL:
        export_excel(rows, out)
        print(f"Excel 报告: {out}")
    out_csv = out.replace(".xlsx", ".csv")
    export_csv(rows, out_csv)
    print(f"CSV 报告: {out_csv}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
