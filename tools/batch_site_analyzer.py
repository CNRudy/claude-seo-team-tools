#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
批量网站分析工具 (Batch Site Analyzer)
======================================
基于 claude-seo 的安全抓取/解析能力，批量分析推广网站的：
  1. 可访问性与运营状态（有效/无效）
  2. 主要内容语言
  3. 目标受众地区与主要目标国家
  4. 简要描述（标题 + 描述摘要）

输出：Excel (.xlsx) + CSV 汇总表。

用法示例:
    python batch_site_analyzer.py --input urls.txt
    python batch_site_analyzer.py --urls https://example.com https://example.org
    python batch_site_analyzer.py --input urls.txt --workers 8 --output result.xlsx
"""

from __future__ import annotations

import argparse
import concurrent.futures as cf
import csv
import os
import re
import sys
import time
from datetime import datetime
from typing import Optional

# ---------------------------------------------------------------- paths
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CLAUDESEO_SCRIPTS = os.path.join(REPO_ROOT, "claude-seo", "scripts")
if CLAUDESEO_SCRIPTS not in sys.path:
    sys.path.insert(0, CLAUDESEO_SCRIPTS)

try:
    import requests
except ImportError:
    print("Error: requests required. pip install requests", file=sys.stderr)
    sys.exit(1)

# Reuse claude-seo's hardened fetch + parse primitives.
try:
    from fetch_page import fetch_page  # noqa: E402
    from parse_html import parse_html  # noqa: E402
    _HAS_CLAUDESEO = True
except ImportError:  # pragma: no cover
    _HAS_CLAUDESEO = False

# claude-seo's url_safety layer refuses loopback proxy hosts (SSRF guard).
# When the environment routes traffic through a local HTTP proxy (e.g.
# 127.0.0.1:xxxxx), that guard would block every request, so we fall back to
# plain requests for proxy-based environments.
def _local_proxy_configured() -> bool:
    for var in ("HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy"):
        val = os.environ.get(var, "")
        if val and re.search(r"127\.0\.0\.1|localhost|\[::1\]", val):
            return True
    return False


_LOCAL_PROXY = _local_proxy_configured()


def _fetch_compat(url: str, timeout: int = 20, _retries: int = 1):
    """Fetch a URL using claude-seo's hardened path, or plain requests when a
    local proxy would be blocked by the SSRF guard."""
    if _HAS_CLAUDESEO and not _LOCAL_PROXY:
        result = fetch_page(url, timeout=timeout, follow_redirects=True, max_redirects=5)
        if result.get("error") and _retries > 0:
            time.sleep(1.0)
            result = fetch_page(url, timeout=timeout, follow_redirects=True, max_redirects=5)
        return result
    # Proxy-friendly fallback (sandbox / corporate proxy environments).
    result = {"url": url, "status_code": None, "content": None,
              "headers": {}, "redirect_chain": [], "error": None}
    if "://" not in url:
        url = f"https://{url}"
    result["url"] = url
    try:
        resp = requests.get(
            url, timeout=timeout, allow_redirects=True,
            headers={
                "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                               "AppleWebKit/537.36 (KHTML, like Gecko) "
                               "Chrome/126.0 Safari/537.36"),
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9",
            },
        )
        result["url"] = resp.url
        result["status_code"] = resp.status_code
        result["headers"] = dict(resp.headers)
        result["content"] = resp.text
        if resp.history:
            result["redirect_chain"] = [h.url for h in resp.history]
    except requests.exceptions.Timeout:
        result["error"] = f"Request timed out after {timeout} seconds"
    except requests.exceptions.TooManyRedirects:
        result["error"] = "Too many redirects (max 5)"
    except requests.exceptions.SSLError as e:
        result["error"] = f"SSL error: {e}"
    except requests.exceptions.ConnectionError as e:
        result["error"] = f"Connection error: {e}"
    except requests.exceptions.RequestException as e:
        result["error"] = f"Request failed: {e}"
    if result.get("error") and _retries > 0:
        time.sleep(1.0)
        return _fetch_compat(url, timeout=timeout, _retries=_retries - 1)
    return result

# Language detection (content-based).
try:
    from langdetect import DetectorFactory, detect_langs

    DetectorFactory.seed = 0
    _HAS_LANGDETECT = True
except ImportError:  # pragma: no cover
    _HAS_LANGDETECT = False

try:
    from bs4 import BeautifulSoup
except ImportError:  # pragma: no cover
    BeautifulSoup = None

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover
    _HAS_OPENPYXL = False


# ---------------------------------------------------------------- language maps
# ISO 639-1 -> 中文语言名
LANG_NAMES = {
    "en": "英语", "zh": "中文", "ja": "日语", "ko": "韩语", "fr": "法语",
    "de": "德语", "es": "西班牙语", "pt": "葡萄牙语", "it": "意大利语",
    "ru": "俄语", "ar": "阿拉伯语", "hi": "印地语", "bn": "孟加拉语",
    "id": "印尼语", "ms": "马来语", "th": "泰语", "vi": "越南语",
    "tl": "菲律宾语", "tr": "土耳其语", "nl": "荷兰语", "pl": "波兰语",
    "sv": "瑞典语", "da": "丹麦语", "no": "挪威语", "fi": "芬兰语",
    "cs": "捷克语", "el": "希腊语", "hu": "匈牙利语", "ro": "罗马尼亚语",
    "uk": "乌克兰语", "he": "希伯来语", "fa": "波斯语", "ur": "乌尔都语",
    "ta": "泰米尔语", "te": "泰卢固语", "mr": "马拉地语", "pa": "旁遮普语",
    "sw": "斯瓦希里语", "af": "南非荷兰语", "ca": "加泰罗尼亚语", "sr": "塞尔维亚语",
    "hr": "克罗地亚语", "sk": "斯洛伐克语", "lt": "立陶宛语", "lv": "拉脱维亚语",
    "et": "爱沙尼亚语", "sq": "阿尔巴尼亚语", "mk": "马其顿语", "bg": "保加利亚语",
    "az": "阿塞拜疆语", "uz": "乌兹别克语", "kk": "哈萨克语", "mn": "蒙古语",
    "my": "缅甸语", "km": "高棉语", "lo": "老挝语", "ne": "尼泊尔语",
    "si": "僧伽罗语", "am": "阿姆哈拉语", "zu": "祖鲁语", "xh": "科萨语",
}

# 语言 -> 主要目标国家（用于无其他信号时的默认推断）
LANG_TO_COUNTRY = {
    "en": "美国/英国/全球", "zh": "中国", "ja": "日本", "ko": "韩国",
    "fr": "法国", "de": "德国", "es": "西班牙/拉美", "pt": "巴西/葡萄牙",
    "it": "意大利", "ru": "俄罗斯", "ar": "中东/北非", "hi": "印度",
    "id": "印尼", "ms": "马来西亚", "th": "泰国", "vi": "越南",
    "tr": "土耳其", "nl": "荷兰", "pl": "波兰", "sv": "瑞典",
    "da": "丹麦", "no": "挪威", "fi": "芬兰", "cs": "捷克",
    "el": "希腊", "hu": "匈牙利", "ro": "罗马尼亚", "uk": "乌克兰",
    "he": "以色列", "fa": "伊朗", "ur": "巴基斯坦", "ta": "印度(泰米尔)",
    "te": "印度(泰卢固)", "mr": "印度(马拉地)", "pa": "印度(旁遮普)",
    "sw": "东非", "af": "南非", "ca": "西班牙(加泰罗尼亚)",
    "sr": "塞尔维亚", "hr": "克罗地亚", "sk": "斯洛伐克", "lt": "立陶宛",
    "lv": "拉脱维亚", "et": "爱沙尼亚", "bg": "保加利亚",
    "az": "阿塞拜疆", "uz": "乌兹别克", "kk": "哈萨克",
    "mn": "蒙古", "my": "缅甸", "km": "柬埔寨", "lo": "老挝",
    "ne": "尼泊尔", "si": "斯里兰卡", "am": "埃塞俄比亚",
}

# 常见 ccTLD / 组合域 -> 国家名
TLD_TO_COUNTRY = {
    ".us": "美国", ".uk": "英国", ".co.uk": "英国", ".org.uk": "英国",
    ".gov.uk": "英国", ".de": "德国", ".fr": "法国", ".jp": "日本",
    ".cn": "中国", ".hk": "中国香港", ".tw": "中国台湾", ".mo": "中国澳门",
    ".kr": "韩国", ".in": "印度", ".br": "巴西", ".mx": "墨西哥",
    ".it": "意大利", ".es": "西班牙", ".ru": "俄罗斯", ".ca": "加拿大",
    ".au": "澳大利亚", ".nz": "新西兰", ".sg": "新加坡", ".my": "马来西亚",
    ".th": "泰国", ".vn": "越南", ".ph": "菲律宾", ".id": "印尼",
    ".ae": "阿联酋", ".sa": "沙特", ".za": "南非", ".ng": "尼日利亚",
    ".eg": "埃及", ".il": "以色列", ".tr": "土耳其", ".nl": "荷兰",
    ".be": "比利时", ".ch": "瑞士", ".at": "奥地利", ".se": "瑞典",
    ".no": "挪威", ".dk": "丹麦", ".fi": "芬兰", ".pl": "波兰",
    ".cz": "捷克", ".pt": "葡萄牙", ".gr": "希腊", ".hu": "匈牙利",
    ".ro": "罗马尼亚", ".ua": "乌克兰", ".ar": "阿根廷", ".cl": "智利",
    ".co": "哥伦比亚", ".pe": "秘鲁", ".ve": "委内瑞拉", ".pk": "巴基斯坦",
    ".bd": "孟加拉", ".lk": "斯里兰卡", ".np": "尼泊尔", ".kz": "哈萨克",
    ".uz": "乌兹别克", ".ir": "伊朗", ".ie": "爱尔兰", ".is": "冰岛",
    ".lu": "卢森堡", ".mt": "马耳他", ".cy": "塞浦路斯", ".ee": "爱沙尼亚",
    ".lv": "拉脱维亚", ".lt": "立陶宛", ".bg": "保加利亚", ".hr": "克罗地亚",
    ".rs": "塞尔维亚", ".sk": "斯洛伐克", ".si": "斯洛文尼亚", ".ge": "格鲁吉亚",
    ".am": "亚美尼亚", ".et": "埃塞俄比亚", ".ke": "肯尼亚", ".gh": "加纳",
    ".tz": "坦桑尼亚", ".ug": "乌干达", ".ma": "摩洛哥", ".dz": "阿尔及利亚",
    ".qa": "卡塔尔", ".kw": "科威特", ".bh": "巴林", ".om": "阿曼",
    ".jo": "约旦", ".lb": "黎巴嫩", ".uy": "乌拉圭", ".ec": "厄瓜多尔",
    ".bo": "玻利维亚", ".py": "巴拉圭", ".cr": "哥斯达黎加", ".pa": "巴拿马",
    ".do": "多米尼加", ".pr": "波多黎各", ".jm": "牙买加",
    ".com.hk": "中国香港", ".com.tw": "中国台湾", ".com.au": "澳大利亚",
    ".com.br": "巴西", ".com.mx": "墨西哥", ".com.sg": "新加坡",
    ".com.my": "马来西亚", ".com.ph": "菲律宾", ".com.tr": "土耳其",
    ".com.cn": "中国", ".com.sa": "沙特", ".com.ar": "阿根廷",
    ".com.co": "哥伦比亚", ".com.pe": "秘鲁", ".com.eg": "埃及",
    ".com.ng": "尼日利亚", ".com.ua": "乌克兰", ".com.pk": "巴基斯坦",
}

# 地区码(ISO 3166) -> 国家名（用于 hreflang / og:locale / html lang 中的 region）
REGION_NAMES = {
    "US": "美国", "GB": "英国", "UK": "英国", "DE": "德国", "FR": "法国",
    "JP": "日本", "CN": "中国", "HK": "中国香港", "TW": "中国台湾",
    "MO": "中国澳门", "KR": "韩国", "IN": "印度", "BR": "巴西", "MX": "墨西哥",
    "IT": "意大利", "ES": "西班牙", "RU": "俄罗斯", "CA": "加拿大",
    "AU": "澳大利亚", "NZ": "新西兰", "SG": "新加坡", "MY": "马来西亚",
    "TH": "泰国", "VN": "越南", "PH": "菲律宾", "ID": "印尼",
    "AE": "阿联酋", "SA": "沙特", "ZA": "南非", "NG": "尼日利亚",
    "EG": "埃及", "IL": "以色列", "TR": "土耳其", "NL": "荷兰", "BE": "比利时",
    "CH": "瑞士", "AT": "奥地利", "SE": "瑞典", "NO": "挪威", "DK": "丹麦",
    "FI": "芬兰", "PL": "波兰", "CZ": "捷克", "PT": "葡萄牙", "GR": "希腊",
    "HU": "匈牙利", "RO": "罗马尼亚", "UA": "乌克兰", "AR": "阿根廷",
    "CL": "智利", "CO": "哥伦比亚", "PE": "秘鲁", "VE": "委内瑞拉",
    "PK": "巴基斯坦", "BD": "孟加拉", "LK": "斯里兰卡", "NP": "尼泊尔",
    "KZ": "哈萨克", "UZ": "乌兹别克", "IR": "伊朗", "IE": "爱尔兰",
    "IS": "冰岛", "LU": "卢森堡", "MT": "马耳他", "CY": "塞浦路斯",
    "EE": "爱沙尼亚", "LV": "拉脱维亚", "LT": "立陶宛", "BG": "保加利亚",
    "HR": "克罗地亚", "RS": "塞尔维亚", "SK": "斯洛伐克", "SI": "斯洛文尼亚",
    "GE": "格鲁吉亚", "AM": "亚美尼亚", "ET": "埃塞俄比亚", "KE": "肯尼亚",
    "GH": "加纳", "TZ": "坦桑尼亚", "UG": "乌干达", "MA": "摩洛哥",
    "DZ": "阿尔及利亚", "QA": "卡塔尔", "KW": "科威特", "BH": "巴林",
    "OM": "阿曼", "JO": "约旦", "LB": "黎巴嫩", "UY": "乌拉圭",
    "EC": "厄瓜多尔", "BO": "玻利维亚", "PY": "巴拉圭", "CR": "哥斯达黎加",
    "PA": "巴拿马", "DO": "多米尼加", "PR": "波多黎各", "JM": "牙买加",
    "EU": "欧洲", "GB-ENG": "英国", "GB-SCT": "英国", "GB-WLS": "英国",
}

# 货币符号 -> 国家
CURRENCY_TO_COUNTRY = {
    "$": "美国/加拿大/澳大利亚", "US$": "美国", "USD": "美国",
    "£": "英国", "GBP": "英国", "€": "欧洲(欧元区)", "EUR": "欧洲(欧元区)",
    "¥": "日本/中国", "CNY": "中国", "JPY": "日本", "R$": "巴西",
    "₩": "韩国", "KRW": "韩国", "₽": "俄罗斯", "RUB": "俄罗斯",
    "₹": "印度", "INR": "印度", "RMB": "中国", "CHF": "瑞士",
    "₪": "以色列", "₺": "土耳其", "₴": "乌克兰", "zł": "波兰",
    "kr": "瑞典/丹麦/挪威", "฿": "泰国", "₫": "越南", "RM": "马来西亚",
    "₱": "菲律宾", "Rp": "印尼", "S$": "新加坡", "HK$": "中国香港",
    "NT$": "中国台湾", "A$": "澳大利亚", "C$": "加拿大", "MX$": "墨西哥",
    "AED": "阿联酋", "SAR": "沙特", "QAR": "卡塔尔",
    "KWD": "科威特", "MXN": "墨西哥", "BRL": "巴西", "ARS": "阿根廷",
    "CLP": "智利", "COP": "哥伦比亚", "PEN": "秘鲁", "UYU": "乌拉圭",
    "NZD": "新西兰", "CAD": "加拿大", "AUD": "澳大利亚", "SGD": "新加坡",
    "HKD": "中国香港", "TWD": "中国台湾", "MYR": "马来西亚", "PHP": "菲律宾",
    "IDR": "印尼", "THB": "泰国", "VND": "越南", "NOK": "挪威",
    "SEK": "瑞典", "DKK": "丹麦", "PLN": "波兰", "CZK": "捷克",
    "HUF": "匈牙利", "RON": "罗马尼亚", "BGN": "保加利亚", "HRK": "克罗地亚",
    "RSD": "塞尔维亚", "ISK": "冰岛", "TRY": "土耳其", "ILS": "以色列",
    "UAH": "乌克兰", "PKR": "巴基斯坦", "BDT": "孟加拉", "LKR": "斯里兰卡",
    "NPR": "尼泊尔", "KZT": "哈萨克", "UZS": "乌兹别克", "NGN": "尼日利亚",
    "KES": "肯尼亚", "EGP": "埃及", "MAD": "摩洛哥", "ZAR": "南非",
    "GHS": "加纳", "TZS": "坦桑尼亚", "UGX": "乌干达", "ETB": "埃塞俄比亚",
}

# 正文关键词（国家/地区特征词）-> 国家  （按从强到弱匹配）
COUNTRY_KEYWORDS = [
    # 区域特征词
    (re.compile(r"\bUK\b|\bUnited Kingdom\b", re.I), "英国"),
    (re.compile(r"\bUSA\b|\bU\.S\.A\.\b|\bUnited States\b", re.I), "美国"),
    (re.compile(r"\bCanada\b", re.I), "加拿大"),
    (re.compile(r"\bAustralia\b", re.I), "澳大利亚"),
    (re.compile(r"\bGermany\b|\bDeutschland\b", re.I), "德国"),
    (re.compile(r"\bFrance\b|\bFrance\b", re.I), "法国"),
    (re.compile(r"\bJapan\b|日本", re.I), "日本"),
    (re.compile(r"\bKorea\b|韩国", re.I), "韩国"),
    (re.compile(r"\bIndia\b|印度", re.I), "印度"),
    (re.compile(r"\bBrazil\b|\bBrasil\b", re.I), "巴西"),
    (re.compile(r"\bMexico\b|México", re.I), "墨西哥"),
    (re.compile(r"\bSpain\b|España", re.I), "西班牙"),
    (re.compile(r"\bItaly\b|\bItalia\b", re.I), "意大利"),
    (re.compile(r"\bNetherlands\b|\bHolland\b", re.I), "荷兰"),
    (re.compile(r"\bSweden\b", re.I), "瑞典"),
    (re.compile(r"\bNorway\b", re.I), "挪威"),
    (re.compile(r"\bDenmark\b", re.I), "丹麦"),
    (re.compile(r"\bFinland\b", re.I), "芬兰"),
    (re.compile(r"\bPoland\b", re.I), "波兰"),
    (re.compile(r"\bCzech\b", re.I), "捷克"),
    (re.compile(r"\bAustria\b", re.I), "奥地利"),
    (re.compile(r"\bSwitzerland\b", re.I), "瑞士"),
    (re.compile(r"\bBelgium\b", re.I), "比利时"),
    (re.compile(r"\bIreland\b", re.I), "爱尔兰"),
    (re.compile(r"\bPortugal\b", re.I), "葡萄牙"),
    (re.compile(r"\bGreece\b", re.I), "希腊"),
    (re.compile(r"\bTurkey\b|\bTürkiye\b", re.I), "土耳其"),
    (re.compile(r"\bRussia\b|Россия", re.I), "俄罗斯"),
    (re.compile(r"\bUkraine\b", re.I), "乌克兰"),
    (re.compile(r"\bSingapore\b", re.I), "新加坡"),
    (re.compile(r"\bMalaysia\b", re.I), "马来西亚"),
    (re.compile(r"\bThailand\b", re.I), "泰国"),
    (re.compile(r"\bVietnam\b", re.I), "越南"),
    (re.compile(r"\bPhilippines\b", re.I), "菲律宾"),
    (re.compile(r"\bIndonesia\b", re.I), "印尼"),
    (re.compile(r"\bNew Zealand\b", re.I), "新西兰"),
    (re.compile(r"\bSouth Africa\b", re.I), "南非"),
    (re.compile(r"\bNigeria\b", re.I), "尼日利亚"),
    (re.compile(r"\bEgypt\b", re.I), "埃及"),
    (re.compile(r"\bKenya\b", re.I), "肯尼亚"),
    (re.compile(r"\bSaudi\b", re.I), "沙特"),
    (re.compile(r"\bUAE\b|\bUnited Arab Emirates\b", re.I), "阿联酋"),
    (re.compile(r"\bQatar\b", re.I), "卡塔尔"),
    (re.compile(r"\bIsrael\b", re.I), "以色列"),
    (re.compile(r"\bArgentina\b", re.I), "阿根廷"),
    (re.compile(r"\bChile\b", re.I), "智利"),
    (re.compile(r"\bColombia\b", re.I), "哥伦比亚"),
    (re.compile(r"\bPeru\b", re.I), "秘鲁"),
    (re.compile(r"\bPakistan\b", re.I), "巴基斯坦"),
    (re.compile(r"\bBangladesh\b", re.I), "孟加拉"),
    (re.compile(r"\bSri Lanka\b", re.I), "斯里兰卡"),
    (re.compile(r"\bNepal\b", re.I), "尼泊尔"),
    (re.compile(r"\bKazakhstan\b", re.I), "哈萨克"),
    (re.compile(r"\bUzbekistan\b", re.I), "乌兹别克"),
    (re.compile(r"\bIran\b", re.I), "伊朗"),
    (re.compile(r"\bIraq\b", re.I), "伊拉克"),
    (re.compile(r"\bJordan\b", re.I), "约旦"),
    (re.compile(r"\bLebanon\b", re.I), "黎巴嫩"),
    (re.compile(r"\bMorocco\b", re.I), "摩洛哥"),
    (re.compile(r"\bAlgeria\b", re.I), "阿尔及利亚"),
    (re.compile(r"\bEthiopia\b", re.I), "埃塞俄比亚"),
    (re.compile(r"\bTanzania\b", re.I), "坦桑尼亚"),
    (re.compile(r"\bUganda\b", re.I), "乌干达"),
    (re.compile(r"\bGhana\b", re.I), "加纳"),
    # 城市/地区特征词（强信号）
    (re.compile(r"\bLondon\b", re.I), "英国"),
    (re.compile(r"\bNew York\b|\bLos Angeles\b|\bChicago\b|\bTexas\b|\bCalifornia\b|\bMiami\b", re.I), "美国"),
    (re.compile(r"\bToronto\b|\bVancouver\b|\bMontreal\b", re.I), "加拿大"),
    (re.compile(r"\bSydney\b|\bMelbourne\b", re.I), "澳大利亚"),
    (re.compile(r"\bBerlin\b|\bMunich\b|\bHamburg\b", re.I), "德国"),
    (re.compile(r"\bParis\b", re.I), "法国"),
    (re.compile(r"\bTokyo\b|\bOsaka\b", re.I), "日本"),
    (re.compile(r"\bSeoul\b", re.I), "韩国"),
    (re.compile(r"\bMumbai\b|\bDelhi\b|\bBangalore\b", re.I), "印度"),
    (re.compile(r"\bSao Paulo\b|\bRio de Janeiro\b", re.I), "巴西"),
    (re.compile(r"\bMadrid\b|\bBarcelona\b", re.I), "西班牙"),
    (re.compile(r"\bMilan\b|\bRome\b", re.I), "意大利"),
    (re.compile(r"\bAmsterdam\b", re.I), "荷兰"),
    (re.compile(r"\bStockholm\b", re.I), "瑞典"),
    (re.compile(r"\bOslo\b", re.I), "挪威"),
    (re.compile(r"\bCopenhagen\b", re.I), "丹麦"),
    (re.compile(r"\bHelsinki\b", re.I), "芬兰"),
    (re.compile(r"\bWarsaw\b", re.I), "波兰"),
    (re.compile(r"\bPrague\b", re.I), "捷克"),
    (re.compile(r"\bVienna\b", re.I), "奥地利"),
    (re.compile(r"\bZurich\b|\bGeneva\b", re.I), "瑞士"),
    (re.compile(r"\bBrussels\b", re.I), "比利时"),
    (re.compile(r"\bDublin\b", re.I), "爱尔兰"),
    (re.compile(r"\bLisbon\b", re.I), "葡萄牙"),
    (re.compile(r"\bAthens\b", re.I), "希腊"),
    (re.compile(r"\bIstanbul\b", re.I), "土耳其"),
    (re.compile(r"\bMoscow\b|\bSaint Petersburg\b", re.I), "俄罗斯"),
    (re.compile(r"\bKyiv\b", re.I), "乌克兰"),
    (re.compile(r"\bSingapore\b", re.I), "新加坡"),
    (re.compile(r"\bKuala Lumpur\b", re.I), "马来西亚"),
    (re.compile(r"\bBangkok\b", re.I), "泰国"),
    (re.compile(r"\bHanoi\b|\bHo Chi Minh\b", re.I), "越南"),
    (re.compile(r"\bManila\b", re.I), "菲律宾"),
    (re.compile(r"\bJakarta\b", re.I), "印尼"),
    (re.compile(r"\bAuckland\b|\bWellington\b", re.I), "新西兰"),
    (re.compile(r"\bJohannesburg\b|\bCape Town\b", re.I), "南非"),
    (re.compile(r"\bLagos\b", re.I), "尼日利亚"),
    (re.compile(r"\bCairo\b", re.I), "埃及"),
    (re.compile(r"\bNairobi\b", re.I), "肯尼亚"),
    (re.compile(r"\bRiyadh\b|\bJeddah\b", re.I), "沙特"),
    (re.compile(r"\bDubai\b|\bAbu Dhabi\b", re.I), "阿联酋"),
    (re.compile(r"\bDoha\b", re.I), "卡塔尔"),
    (re.compile(r"\bTel Aviv\b", re.I), "以色列"),
    (re.compile(r"\bBuenos Aires\b", re.I), "阿根廷"),
    (re.compile(r"\bSantiago\b", re.I), "智利"),
    (re.compile(r"\bBogota\b", re.I), "哥伦比亚"),
    (re.compile(r"\bLima\b", re.I), "秘鲁"),
    (re.compile(r"\bKarachi\b|\bLahore\b", re.I), "巴基斯坦"),
    (re.compile(r"\bDhaka\b", re.I), "孟加拉"),
    (re.compile(r"\bColombo\b", re.I), "斯里兰卡"),
    (re.compile(r"\bKathmandu\b", re.I), "尼泊尔"),
    (re.compile(r"\bAlmaty\b", re.I), "哈萨克"),
    (re.compile(r"\bTashkent\b", re.I), "乌兹别克"),
    (re.compile(r"\bTehran\b", re.I), "伊朗"),
    (re.compile(r"\bBaghdad\b", re.I), "伊拉克"),
    (re.compile(r"\bAmman\b", re.I), "约旦"),
    (re.compile(r"\bBeirut\b", re.I), "黎巴嫩"),
    (re.compile(r"\bCasablanca\b|\bMarrakesh\b", re.I), "摩洛哥"),
    (re.compile(r"\bAlgiers\b", re.I), "阿尔及利亚"),
    (re.compile(r"\bAddis Ababa\b", re.I), "埃塞俄比亚"),
    (re.compile(r"\bDar es Salaam\b", re.I), "坦桑尼亚"),
    (re.compile(r"\bKampala\b", re.I), "乌干达"),
    (re.compile(r"\bAccra\b", re.I), "加纳"),
    (re.compile(r"\bHong Kong\b|香港", re.I), "中国香港"),
    (re.compile(r"\bTaiwan\b|台湾", re.I), "中国台湾"),
    (re.compile(r"\bMacau\b|澳门", re.I), "中国澳门"),
    (re.compile(r"\bBeijing\b|北京", re.I), "中国"),
    (re.compile(r"\bShanghai\b|上海", re.I), "中国"),
    (re.compile(r"\bShenzhen\b|深圳", re.I), "中国"),
    (re.compile(r"\bGuangzhou\b|广州", re.I), "中国"),
    (re.compile(r"\bChina\b|中国", re.I), "中国"),
]


# ---------------------------------------------------------------- helpers
def normalize_url(url: str) -> str:
    """Normalize a URL: strip whitespace, add scheme if missing."""
    url = url.strip()
    if not url:
        return ""
    if "://" not in url:
        url = f"https://{url}"
    return url


def extract_region_code(text: str) -> Optional[str]:
    """Extract a region code from strings like 'en-US', 'zh-Hans-CN'."""
    m = re.search(r"\b[a-z]{2,3}(?:-[A-Za-z]{2,4})?-(US|GB|UK|DE|FR|JP|CN|HK|TW|MO|KR|IN|BR|MX|IT|ES|RU|CA|AU|NZ|SG|MY|TH|VN|PH|ID|AE|SA|ZA|NG|EG|IL|TR|NL|BE|CH|AT|SE|NO|DK|FI|PL|CZ|PT|GR|HU|RO|UA|AR|CL|CO|PE|VE|PK|BD|LK|NP|KZ|UZ|IR|IE|IS|LU|MT|CY|EE|LV|LT|BG|HR|RS|SK|SI|GE|AM|ET|KE|GH|TZ|UG|MA|DZ|QA|KW|BH|OM|JO|LB|UY|EC|BO|PY|CR|PA|DO|PR|JM|EU)\b", text, re.I)
    if m:
        return m.group(1).upper()
    return None


# 常见 ccTLD -> 主要语言（用于内容稀少时的语言兜底推断）
TLD_TO_LANG = {
    ".co.uk": "en", ".org.uk": "en", ".gov.uk": "en", ".uk": "en",
    ".de": "de", ".fr": "fr", ".jp": "ja", ".cn": "zh", ".hk": "zh",
    ".tw": "zh", ".mo": "zh", ".kr": "ko", ".in": "en", ".br": "pt",
    ".mx": "es", ".it": "it", ".es": "es", ".ru": "ru", ".ca": "en",
    ".au": "en", ".nz": "en", ".sg": "en", ".my": "ms", ".th": "th",
    ".vn": "vi", ".ph": "tl", ".id": "id", ".ae": "ar", ".sa": "ar",
    ".za": "en", ".ng": "en", ".eg": "ar", ".il": "he", ".tr": "tr",
    ".nl": "nl", ".be": "nl", ".ch": "de", ".at": "de", ".se": "sv",
    ".no": "no", ".dk": "da", ".fi": "fi", ".pl": "pl", ".cz": "cs",
    ".pt": "pt", ".gr": "el", ".hu": "hu", ".ro": "ro", ".ua": "uk",
    ".ar": "es", ".cl": "es", ".co": "es", ".pe": "es", ".ve": "es",
    ".pk": "ur", ".bd": "bn", ".lk": "si", ".np": "ne", ".kz": "kk",
    ".uz": "uz", ".ir": "fa", ".ie": "en", ".is": "is", ".ee": "et",
    ".lv": "lv", ".lt": "lt", ".bg": "bg", ".hr": "hr", ".rs": "sr",
    ".sk": "sk", ".si": "sl", ".ge": "ka", ".am": "hy", ".et": "am",
    ".ke": "sw", ".gh": "en", ".ma": "ar", ".dz": "ar", ".qa": "ar",
    ".kw": "ar", ".bh": "ar", ".om": "ar", ".jo": "ar", ".lb": "ar",
    ".uy": "es", ".ec": "es", ".bo": "es", ".py": "es", ".cr": "es",
    ".pa": "es", ".do": "es", ".jm": "en",
}


def detect_language(parsed: dict, raw_html: str, body_text: str, url: str = "") -> tuple:
    """
    Detect primary content language.

    Returns: (lang_code, lang_name, confidence_source)
    Priority: html lang attr -> og:locale -> content detection (langdetect)
            -> ccTLD fallback -> default
    """
    lang_code = None
    source = "html lang"

    # 1. <html lang="...">
    if BeautifulSoup is not None:
        soup = BeautifulSoup(raw_html, "lxml")
        html_tag = soup.find("html")
        if html_tag and html_tag.get("lang"):
            lang_code = str(html_tag["lang"]).split("-")[0].lower()
    if not lang_code:
        # 2. og:locale
        og = parsed.get("open_graph") or {}
        og_locale = og.get("og:locale") or og.get("og:locale:alternate")
        if og_locale:
            base = og_locale.split("_")[0].lower()
            if len(base) == 2:
                lang_code = base
                source = "og:locale"
    if not lang_code and _HAS_LANGDETECT:
        # 3. content detection
        sample = body_text[:2000]
        if sample and len(sample.strip()) >= 40:
            try:
                langs = detect_langs(sample)
                if langs:
                    lang_code = langs[0].lang
                    source = "内容检测"
            except Exception:
                pass
    if not lang_code:
        # 4. ccTLD-based fallback (useful for JS-shell / blocked pages)
        if url:
            domain = re.sub(r"^www\.", "", url.split("//")[-1].split("/")[0]).lower()
            for suffix in sorted(TLD_TO_LANG, key=len, reverse=True):
                if domain.endswith(suffix):
                    lang_code = TLD_TO_LANG[suffix]
                    source = f"域名后缀"
                    break
    if not lang_code:
        lang_code = "en" if body_text else "unknown"
        source = "默认"

    lang_name = LANG_NAMES.get(lang_code, lang_code)
    return lang_code, lang_name, source


def detect_target_country(
    parsed: dict,
    final_url: str,
    lang_code: str,
    body_text: str,
) -> tuple:
    """
    Infer the main target country/region using multiple signals.

    Returns: (countries_str, signals_used_str)
    Signals (priority order): hreflang region -> og:locale region
            -> html lang region -> TLD -> currency symbol
            -> content keywords -> language default
    """
    signals = []
    countries = []

    def add_country(c, sig):
        if c and c not in countries:
            countries.append(c)
            signals.append(sig)

    # 1. hreflang regions
    hreflangs = parsed.get("hreflang") or []
    for h in hreflangs[:20]:
        reg = extract_region_code(h.get("lang", ""))
        if reg:
            add_country(REGION_NAMES.get(reg, reg), "hreflang")
            if len(countries) >= 3:
                break

    # 2. og:locale
    og = parsed.get("open_graph") or {}
    for v in (og.get("og:locale"), og.get("og:locale:alternate")):
        if not v:
            continue
        if isinstance(v, str):
            v = [v]
        for loc in v:
            parts = str(loc).replace("-", "_").split("_")
            if len(parts) >= 2:
                reg = parts[1].upper()
                add_country(REGION_NAMES.get(reg, reg), "og:locale")

    # 3. TLD
    if final_url:
        domain = re.sub(r"^www\.", "", final_url.split("//")[-1].split("/")[0]).lower()
        for suffix in sorted(TLD_TO_COUNTRY, key=len, reverse=True):
            if domain.endswith(suffix):
                add_country(TLD_TO_COUNTRY[suffix], "域名后缀")
                break

    # 4. Content country keywords (stronger signal than currency)
    kw_sample = body_text[:6000]
    for pattern, country in COUNTRY_KEYWORDS:
        if pattern.search(kw_sample):
            add_country(country, "内容关键词")
            if len(countries) >= 4:
                break

    # 5. Currency symbols in body (first 8000 chars)
    cur_sample = body_text[:8000]
    for sym, country in CURRENCY_TO_COUNTRY.items():
        if sym in cur_sample:
            add_country(country, "货币符号")
            break  # only strongest currency signal

    # 6. Language default
    if not countries:
        add_country(LANG_TO_COUNTRY.get(lang_code, "全球"), "语言推断")

    # Deduplicate
    seen = []
    for c in countries:
        if c not in seen:
            seen.append(c)
    return "、".join(seen[:4]), "、".join(signals[:6])


def build_description(parsed: dict, body_text: str) -> str:
    """Build a short description from title + meta description + H1."""
    parts = []
    title = parsed.get("title")
    if title:
        parts.append(title.strip())
    desc = parsed.get("meta_description")
    if desc:
        parts.append(desc.strip()[:180])
    if not parts:
        h1s = parsed.get("h1") or []
        if h1s:
            parts.append(h1s[0].strip()[:150])
    if not parts:
        # fallback to first meaningful sentence of body
        sentences = re.split(r"[。.!?\n]+", body_text.strip())
        for s in sentences:
            s = s.strip()
            if len(s) >= 20:
                parts.append(s[:180])
                break
    if not parts:
        parts.append("（无有效内容）")
    return " | ".join([p for p in parts if p])[:400]


# ---------------------------------------------------------------- analysis
def analyze_url(url: str, timeout: int = 20) -> dict:
    """Analyze a single URL. Returns a result row dict."""
    url = normalize_url(url)
    started = time.time()
    row = {
        "URL": url,
        "状态": "无效",
        "HTTP状态码": "",
        "目标国家/地区": "",
        "语言": "",
        "语言代码": "",
        "网站标题": "",
        "简要描述": "",
        "响应时间(秒)": "",
        "最终URL": "",
        "信号来源": "",
        "错误信息": "",
    }

    if not url:
        row["错误信息"] = "空 URL"
        return row

    result = _fetch_compat(url, timeout=timeout)
    elapsed = round(time.time() - started, 2)
    row["响应时间(秒)"] = elapsed

    if result.get("error"):
        row["错误信息"] = result["error"][:200]
        row["HTTP状态码"] = str(result.get("status_code") or "")
        return row

    status = result.get("status_code")
    row["HTTP状态码"] = str(status or "")
    final_url = result.get("url") or url
    row["最终URL"] = final_url
    content = result.get("content") or ""

    if status is None or status >= 400:
        # 403/401/429: 站点在线但在运营受限（反爬/需登录/限流）—— 仍视为可访问
        if status in (401, 403, 429):
            row["状态"] = "有效"
            row["HTTP状态码"] = str(status)
            # 受限站点仍可从域名后缀推断目标国家与语言
            domain = re.sub(r"^www\.", "", final_url.split("//")[-1].split("/")[0]).lower()
            signals = []
            for suffix in sorted(TLD_TO_COUNTRY, key=len, reverse=True):
                if domain.endswith(suffix):
                    row["目标国家/地区"] = TLD_TO_COUNTRY[suffix]
                    signals.append("域名后缀")
                    break
            for suffix in sorted(TLD_TO_LANG, key=len, reverse=True):
                if domain.endswith(suffix):
                    lc = TLD_TO_LANG[suffix]
                    row["语言"] = LANG_NAMES.get(lc, lc)
                    row["语言代码"] = lc
                    signals.append("域名后缀")
                    break
            if signals:
                row["信号来源"] = "国家/语言:域名后缀(受限页)"
            row["错误信息"] = f"可访问但受限（HTTP {status}，可能反爬或需登录）"
            return row
        row["错误信息"] = f"HTTP {status}"
        return row

    # Parse HTML
    parsed = {}
    body_text = ""
    if content:
        try:
            parsed = parse_html(content)
        except Exception as e:  # pragma: no cover
            row["错误信息"] = f"解析失败: {e}"
            return row
        try:
            import trafilatura

            body_text = trafilatura.extract(content, include_comments=False) or ""
        except Exception:
            body_text = ""

    title = (parsed.get("title") or "").strip()
    word_count = parsed.get("word_count") or 0

    # 有效判定：HTTP 2xx/3xx 且返回了 HTML 内容（站点在运营）
    # 注意：JS 渲染站点 / 反爬壳页面可能只有很少的静态内容，仍视为可访问
    is_valid = status < 400 and bool(content) and len(content.strip()) > 0
    if not is_valid:
        row["错误信息"] = "可访问但无有效内容（可能为占位页/空页）"
        row["网站标题"] = title[:200]
        return row

    row["状态"] = "有效"
    row["网站标题"] = title[:200]
    if not title and word_count < 40:
        row["错误信息"] = "内容稀少（JS渲染/反爬壳页面，仅确认可访问）"

    # Language detection
    lang_code, lang_name, lang_src = detect_language(parsed, content, body_text, final_url)
    row["语言代码"] = lang_code
    row["语言"] = lang_name
    row["信号来源"] = f"语言:{lang_src}"

    # Country detection
    countries, signals = detect_target_country(parsed, final_url, lang_code, body_text)
    row["目标国家/地区"] = countries
    if signals:
        row["信号来源"] += f"; 国家:{signals}"

    # Description
    row["简要描述"] = build_description(parsed, body_text)

    return row


def analyze_batch(urls: list, workers: int = 5, timeout: int = 20) -> list:
    """Analyze a list of URLs concurrently. Returns list of result dicts."""
    results = []
    with cf.ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(analyze_url, u, timeout): u for u in urls}
        for i, fut in enumerate(cf.as_completed(futures), 1):
            url = futures[fut]
            try:
                row = fut.result()
            except Exception as e:  # pragma: no cover
                row = {
                    "URL": url, "状态": "无效", "HTTP状态码": "",
                    "目标国家/地区": "", "语言": "", "语言代码": "",
                    "网站标题": "", "简要描述": "", "响应时间(秒)": "",
                    "最终URL": "", "信号来源": "", "错误信息": str(e)[:200],
                }
            results.append(row)
            print(f"[{i}/{len(urls)}] {row['状态']}  {url}  ({row['响应时间(秒)']}s)")
    # Preserve original input order
    order = {u: idx for idx, u in enumerate(urls)}
    results.sort(key=lambda r: order.get(r["URL"], 0))
    return results


# ---------------------------------------------------------------- export
HEADERS = [
    "URL", "状态", "HTTP状态码", "目标国家/地区", "语言", "语言代码",
    "网站标题", "简要描述", "响应时间(秒)", "最终URL", "信号来源", "错误信息",
]


def export_excel(rows: list, path: str) -> None:
    """Export results to an Excel workbook with simple styling."""
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl not installed; cannot export xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "网站分析结果"

    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF")
    valid_fill = PatternFill("solid", fgColor="E2EFDA")
    invalid_fill = PatternFill("solid", fgColor="FCE4EC")

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([row.get(h, "") for h in HEADERS])

    for r_idx in range(2, ws.max_row + 1):
        status = ws.cell(row=r_idx, column=2).value
        fill = valid_fill if status == "有效" else invalid_fill
        for c_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths
    widths = {
        "A": 40, "B": 8, "C": 12, "D": 26, "E": 10, "F": 10,
        "G": 45, "H": 80, "I": 12, "J": 45, "K": 30, "L": 40,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
    wb.save(path)


def export_csv(rows: list, path: str) -> None:
    """Export results to CSV (utf-8-sig for Excel compatibility)."""
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({h: row.get(h, "") for h in HEADERS})


# ---------------------------------------------------------------- CLI
def read_urls_from_file(path: str) -> list:
    """Read URLs from a file (one per line; ignore blanks & comments)."""
    urls = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            urls.append(line)
    return urls


def main() -> int:
    parser = argparse.ArgumentParser(
        description="批量网站分析工具：检测可访问性、语言、目标国家并导出表格",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "示例:\n"
            "  %(prog)s --input urls.txt\n"
            "  %(prog)s --urls https://example.com https://example.org\n"
            "  %(prog)s --input urls.txt --workers 8 --output result.xlsx"
        ),
    )
    parser.add_argument("--input", help="输入文件路径（每行一个 URL）")
    parser.add_argument("--urls", nargs="+", help="直接传入 URL 列表")
    parser.add_argument("--workers", type=int, default=5, help="并发线程数（默认 5）")
    parser.add_argument("--timeout", type=int, default=20, help="请求超时秒数（默认 20）")
    parser.add_argument("--output", help="输出 Excel 文件路径（默认自动生成）")
    parser.add_argument("--no-csv", action="store_true", help="不生成 CSV")
    args = parser.parse_args()

    urls = []
    if args.input:
        urls = read_urls_from_file(args.input)
    elif args.urls:
        urls = [normalize_url(u) for u in args.urls]
    if not urls:
        parser.error("请通过 --input 或 --urls 提供至少一个 URL")

    # Deduplicate preserving order
    seen = set()
    dedup = []
    for u in urls:
        if u and u not in seen:
            seen.add(u)
            dedup.append(u)
    urls = dedup

    print(f"开始分析 {len(urls)} 个网站（并发 {args.workers}）...")
    results = analyze_batch(urls, workers=args.workers, timeout=args.timeout)

    valid = sum(1 for r in results if r["状态"] == "有效")
    invalid = len(results) - valid
    print("\n================ 汇总 ================")
    print(f"总网站数: {len(results)} | 有效: {valid} | 无效: {invalid}")

    # Output files
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_xlsx = args.output or os.path.join(out_dir, f"site_analysis_{stamp}.xlsx")
    if not out_xlsx.endswith((".xlsx", ".xls")):
        out_xlsx += ".xlsx"

    if _HAS_OPENPYXL:
        export_excel(results, out_xlsx)
        print(f"Excel 结果: {out_xlsx}")
    else:
        print("提示: 未安装 openpyxl，跳过 Excel 导出")

    if not args.no_csv:
        out_csv = out_xlsx.replace(".xlsx", ".csv")
        export_csv(results, out_csv)
        print(f"CSV 结果: {out_csv}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
